import os
import shutil
from openpyxl import load_workbook, Workbook


def find_excel_and_copy_rename(directory, level):
    # Check if the specified directory exists
    if not os.path.isdir(directory):
        print("Error: The specified directory does not exist.")
        return
    
    # Iterate through files in the directory
    for filename in os.listdir(directory):
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            if level in filename:
                # Construct the source and destination paths
                source_path = os.path.join(directory, filename)
                new_filename = f"{part_number} Logistics"  # New filename
                destination_path = os.path.join(directory, new_filename)
                
                # Copy and rename the file
                try:
                    shutil.copy(source_path, destination_path)
                    print(f"Excel file '{filename}' copied and renamed to '{new_filename}'.")
                except Exception as e:
                    print(f"An error occurred while copying and renaming the file '{filename}': {e}")

# Example usage:
specified_directory = input("Template Folder Path: ")  # Specify the directory path
level = input("Managed Level: ")  # User input
find_excel_and_copy_rename(specified_directory, level)


def auto_fill_excel_existing(directory, level, existing_file, sheet_name1, sheet_name2, sheet_name3, part_numbers):
    # Load the existing workbook
    sheet_name1= "Request Details"
    sheet_name2= "General Details"
    sheet_name3= "Image"

    existing_file = find_excel_and_copy_rename(directory, level)

 #   list_of_part_numbers = part_numbers.toPlainText().splitlines()
  #      for line in list_of_part_numbers:
   #         line = line.strip()
   #         part_number = line

    try:
        wb = load_workbook(existing_file)
    except FileNotFoundError:
        print("Error: The specified Excel file does not exist.")
        return
    
    # Check if the target sheet exists, otherwise create it
    if sheet_name1 in wb.sheetnames:
        ws1 = wb[sheet_name1]
    else:
        print(f"Error: The sheet '{sheet_name1}' does not exist in the Excel file.")
        return

    if sheet_name2 in wb.sheetnames:
        ws2 = wb[sheet_name2]
     else:
        print(f"Error: The sheet '{sheet_name2}' does not exist in the Excel file.")
        return

    if sheet_name2 in wb.sheetnames:
        ws2 = wb[sheet_name3]
     else:
        print(f"Error: The sheet '{sheet_name3}' does not exist in the Excel file.")
        return

    # Get inputs from the user
    Date = input("Enter the Date: ")
    Requestor = input("Enter the Requestor: ")
    EPC = input("Enter the Engineering Point of Contact: ")
    SBU = input("Enter the SBU: ")
    Part_Number = input("Enter the Part Number(s): ")
    Image = input("Upload the image(s): ")
    
    # Fill cells with inputs
    ws1['A2'] = Date
    ws1['B2'] = Requestor
    ws1['C2'] = EPC
    ws1['D2'] = SBU
    ws2['A2'] = Part_Number
    ws3['B2'] = Image
    
    # Save the changes to the Excel file
    wb.save(existing_file)
    print("Excel file updated successfully.")

# Call the function to execute
existing_excel_file = find_excel_and_copy_rename(directory, level)  # Specify the existing Excel file path
sheet_name1= "Request Details"
sheet_name2= "General Details"
sheet_name3= "Image"
auto_fill_excel_existing(existing_excel_file, sheet_name1, sheet_name2, sheet_name3)
